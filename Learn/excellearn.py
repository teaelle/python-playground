import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 66
# rows = [
#     (88, 46),
#     (89, 38),
#     (23, 59),
#     (56, 21),
#     (24, 18),
#     (34, 15)
# ]
# for row in rows:
#     ws.append(row)

# ws['A20'] = 42
# ws['B1'] = 42
wb.save("sample.xlsx")